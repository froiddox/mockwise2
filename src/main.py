from collections.abc import Callable

import openpyxl

from inserter import execute_test_case
from plugin_loader import load_plugins
from reader import read_excel
from screen_matcher import find_reference_on_screen
from ui import App


def _mark_done(file_path: str, plugin_key: str, row_index: int) -> None:
    """Write 'Done' to the Status column of the given Excel row.
    Creates the Status column header if it does not yet exist.
    """
    wb = openpyxl.load_workbook(file_path)
    ws = next(s for s in wb.worksheets if s.title.strip().lower() == plugin_key)

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    status_col: int | None = None
    for cell in header_row:
        if cell.value and str(cell.value).strip().lower() == "status":
            status_col = cell.column
            break

    if status_col is None:
        last_col = max((c.column for c in header_row if c.value is not None), default=0)
        status_col = last_col + 1
        ws.cell(row=1, column=status_col, value="Status")

    ws.cell(row=row_index, column=status_col, value="Done")
    wb.save(file_path)


def run_import(
    file_path: str,
    status: Callable[[str], None],
    progress: Callable[[int, int], None],
) -> None:
    plugins = load_plugins()
    sheets = read_excel(file_path)

    if not sheets:
        status("No sheets found in workbook.")
        return

    total = sum(
        len(cases)
        for sheet_name, cases in sheets
        if plugins.get(sheet_name) is not None
    )
    done = 0
    progress(done, total)

    for sheet_name, test_cases in sheets:
        plugin = plugins.get(sheet_name)
        if plugin is None:
            status(f"Skipped '{sheet_name}' — no matching plugin.")
            continue

        n = len(test_cases)
        width = len(str(n))
        status(f"Plugin: {plugin.name}  ({n} cases)")

        for i, tc in enumerate(test_cases, start=1):
            try:
                anchor, _ = find_reference_on_screen(plugin.image_path)
            except RuntimeError as exc:
                status(f"  Error: {exc}")
                status("Run aborted.")
                return

            execute_test_case(plugin, tc, anchor)

            try:
                _mark_done(file_path, sheet_name, tc.row_index)
            except Exception as exc:
                status(f"  Warning: could not update Excel status: {exc}")

            done += 1
            progress(done, total)
            status(f"  [{i:{width}}/{n}] done.")

    status(f"Finished — {done} case(s) processed.")


def main() -> None:
    app = App(on_import=run_import)
    app.mainloop()


if __name__ == "__main__":
    main()
